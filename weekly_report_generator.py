"""
Weekly Market Report Generator
--------------------------------
Pulls recent price data for a list of tickers (stocks, ETFs, or bond proxies)
and builds a formatted Excel report with:
  - A summary table (price, weekly % change, volume) with conditional formatting
  - A price trend chart per ticker
  - Auto-sized columns and a clean, professional layout

Usage:
    python weekly_report_generator.py AAPL MSFT TLT GLD

Requires:
    pip install yfinance openpyxl pandas
"""

import sys
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter


# ---------- Styling constants ----------
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F2937")
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(color="006100")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")


# ---------- Data fetching ----------
def fetch_data(tickers, lookback_days=30):
    """Pull daily price history for each ticker via yfinance.

    Returns a dict: {ticker: pandas.DataFrame with Date, Close, Volume}
    """
    import yfinance as yf  # imported here so the rest of the module works without it installed

    end = datetime.today()
    start = end - timedelta(days=lookback_days)

    data = {}
    for ticker in tickers:
        hist = yf.Ticker(ticker).history(start=start, end=end)
        if hist.empty:
            print(f"Warning: no data returned for {ticker}, skipping.")
            continue
        hist = hist.reset_index()[["Date", "Close", "Volume"]]
        data[ticker] = hist
    return data


# ---------- Report building (pure openpyxl, no network needed) ----------
def build_report(data, output_path="weekly_report.xlsx"):
    """Build the formatted Excel report from a dict of {ticker: DataFrame}."""
    wb = Workbook()

    # ---- Summary sheet ----
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = f"Weekly Market Report — {datetime.today().strftime('%B %d, %Y')}"
    summary["A1"].font = TITLE_FONT
    summary.merge_cells("A1:E1")

    headers = ["Ticker", "Last Close", "Weekly % Change", "Avg Volume", "Trend"]
    for col, h in enumerate(headers, start=1):
        cell = summary.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row = 4
    for ticker, df in data.items():
        if len(df) < 2:
            continue
        last_close = df["Close"].iloc[-1]
        first_close = df["Close"].iloc[0]
        pct_change = (last_close - first_close) / first_close
        avg_volume = df["Volume"].mean()

        summary.cell(row=row, column=1, value=ticker).font = BODY_FONT
        summary.cell(row=row, column=2, value=round(last_close, 2)).number_format = "$#,##0.00"
        pct_cell = summary.cell(row=row, column=3, value=pct_change)
        pct_cell.number_format = "0.00%"
        summary.cell(row=row, column=4, value=int(avg_volume)).number_format = "#,##0"
        summary.cell(row=row, column=5, value=f"See '{ticker}' tab for chart")

        for c in range(1, 6):
            summary.cell(row=row, column=c).border = THIN_BORDER

        # Build a per-ticker sheet with the raw data and a trend chart
        _build_ticker_sheet(wb, ticker, df)
        row += 1

    # Conditional formatting: green if weekly change >= 0, red if negative
    last_row = row - 1
    if last_row >= 4:
        rng = f"C4:C{last_row}"
        summary.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT)
        )
        summary.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT)
        )

    # Column widths
    widths = [12, 14, 18, 16, 30]
    for i, w in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"Report saved to {output_path}")


def _build_ticker_sheet(wb, ticker, df):
    ws = wb.create_sheet(title=ticker[:31])  # Excel sheet name limit is 31 chars

    ws["A1"] = f"{ticker} — Price History"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    ws.append([])  # spacer
    ws.append(["Date", "Close", "Volume"])
    for c in range(1, 4):
        cell = ws.cell(row=3, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for _, r in df.iterrows():
        ws.append([r["Date"].strftime("%Y-%m-%d"), round(r["Close"], 2), int(r["Volume"])])

    # Line chart of closing price
    chart = LineChart()
    chart.title = f"{ticker} Closing Price"
    chart.y_axis.title = "Price ($)"
    chart.x_axis.title = "Date"
    chart.height = 8
    chart.width = 18

    n = len(df)
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + n)
    cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "E3")

    for i, w in enumerate([12, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- CLI entry point ----------
if __name__ == "__main__":
    tickers = sys.argv[1:] or ["AAPL", "MSFT", "TLT", "GLD"]
    print(f"Fetching data for: {', '.join(tickers)}")
    data = fetch_data(tickers)
    build_report(data)
